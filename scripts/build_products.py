#!/usr/bin/env python3
"""Build the product catalog data + download product images.

Reads `public/260717 _ Danh sách sản phẩm An Thái.xlsx`, then:
  1. Extracts the 3-level hierarchy (Bộ phận → Nhóm mẹ → Sản phẩm con).
  2. Downloads the Google Drive images linked from the cells:
       - column E (Nhóm)  -> parent-group image  (first row of each group)
       - column F (Mã AT) -> individual product image
     into `public/product/`, skipping files that already exist.
  3. Regenerates `src/data/products.ts` with local `image` paths.

Idempotent: re-run any time to fill in images that failed / were added.
Requires: openpyxl (already a dev dependency of the extraction step).

Usage:
    python scripts/build_products.py            # download missing + rebuild TS
    python scripts/build_products.py --no-download   # rebuild TS only
"""
import json
import os
import re
import sys
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed

import io
import urllib.request
import urllib.error

import openpyxl
from PIL import Image, ImageOps

# Product cards render ~300px, the detail hero ~600px — 1000px is plenty for
# retina while keeping the shipped bundle small.
MAX_DIM = 1000
JPEG_QUALITY = 82

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'public', '260717 _ Danh sách sản phẩm An Thái.xlsx')
IMG_DIR = os.path.join(ROOT, 'public', 'product')
TS_OUT = os.path.join(ROOT, 'src', 'data', 'products.ts')
SHEETS = ['VN AT - XP - MAT', 'VN XCBB']

CAT_SLUGS = {
    'Động cơ tổng thành & Phụ kiện': 'dong-co',
    'Hệ thống khung gầm': 'khung-gam',
    'Hệ thống phanh': 'phanh',
    'Hệ thống cabin': 'cabin',
}
CAT_ORDER = [
    'Động cơ tổng thành & Phụ kiện',
    'Hệ thống khung gầm',
    'Hệ thống phanh',
    'Hệ thống cabin',
]
BRAND_ORDER = ['Antek', 'XCBB', 'MAT', 'X-POWER']


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def norm_brand(b):
    if not b:
        return 'Antek'
    b = str(b).strip()
    u = b.upper()
    if u.startswith('ANTEK'):
        return 'Antek'
    if u.startswith('XCBB'):
        return 'XCBB'
    if u.startswith('X-POWER') or u.startswith('XPOWER'):
        return 'X-POWER'
    if u.startswith('MAT'):
        return 'MAT'
    return b


def slugify(s):
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd').replace('Đ', 'D').lower()
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s or 'nhom'


def split_lines(v):
    if not v:
        return []
    out = []
    for line in str(v).split('\n'):
        line = re.sub(r'^[-•–]\s*', '', line.strip()).strip()
        if line:
            out.append(line)
    return out


def drive_id(url):
    if not url:
        return None
    m = re.search(r'[?&]id=([\w-]+)', url) or re.search(r'/d/([\w-]+)', url)
    return m.group(1) if m else None


# --------------------------------------------------------------------------- #
# Extraction
# --------------------------------------------------------------------------- #
def extract():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    cats = {}  # catname -> {slug,name,groups:{gname:{slug,name,imageId,products:[]}}}
    used_ids = set()
    order = 0

    def make_id(code):
        base = re.sub(r'[^A-Za-z0-9]+', '-', str(code)).strip('-').lower() if code else f'sp-{order}'
        pid, n = base, 2
        while pid in used_ids:
            pid = f'{base}-{n}'
            n += 1
        used_ids.add(pid)
        return pid

    for sname in SHEETS:
        ws = wb[sname]
        cur_bp = cur_nhom = None
        for r in range(3, ws.max_row + 1):
            bp = ws.cell(r, 3).value
            nhom = ws.cell(r, 5).value
            code = ws.cell(r, 6).value
            name_old = ws.cell(r, 7).value
            name_new = ws.cell(r, 8).value
            unit = ws.cell(r, 9).value
            vehicles = ws.cell(r, 10).value
            specs = ws.cell(r, 11).value
            intro = ws.cell(r, 12).value
            brand = ws.cell(r, 13).value
            group_link = ws.cell(r, 5).hyperlink.target if ws.cell(r, 5).hyperlink else None
            prod_link = ws.cell(r, 6).hyperlink.target if ws.cell(r, 6).hyperlink else None

            if bp:
                cur_bp = str(bp).strip()
            if nhom:
                cur_nhom = str(nhom).strip()
            if not code and not name_old and not name_new:
                continue
            if cur_bp not in CAT_SLUGS:
                continue

            cslug = CAT_SLUGS[cur_bp]
            cat = cats.setdefault(cur_bp, {'slug': cslug, 'name': cur_bp, 'groups': {}})
            gname = cur_nhom or 'Khác'
            grp = cat['groups'].get(gname)
            if grp is None:
                grp = {'slug': None, 'name': gname, 'imageId': None, 'products': []}
                cat['groups'][gname] = grp
            # First group image link encountered wins (merged brands share a card).
            if grp['imageId'] is None and drive_id(group_link):
                grp['imageId'] = drive_id(group_link)

            order += 1
            pid = make_id(code)
            nm = (str(name_new).strip() if name_new else '') or (str(name_old).strip() if name_old else '')
            grp['products'].append({
                'id': pid,
                'code': str(code).strip() if code else '',
                'name': nm,
                'fullName': str(name_old).strip() if name_old else nm,
                'brand': norm_brand(brand),
                'unit': str(unit).strip() if unit else '',
                'vehicles': str(vehicles).strip() if vehicles else '',
                'specs': split_lines(specs),
                'intro': split_lines(intro),
                'categorySlug': cslug,
                'groupSlug': None,
                'imageId': drive_id(prod_link),
                'image': '',
            })

    # Assign group slugs (unique within category).
    for cat in cats.values():
        seen = set()
        for gname, grp in cat['groups'].items():
            s = slugify(gname)
            base, n = s, 2
            while s in seen:
                s = f'{base}-{n}'
                n += 1
            seen.add(s)
            grp['slug'] = s
            grp['image'] = ''
            for p in grp['products']:
                p['groupSlug'] = s

    ordered = [cats[c] for c in CAT_ORDER if c in cats]
    return ordered


# --------------------------------------------------------------------------- #
# Image download
# --------------------------------------------------------------------------- #
EXT_BY_MAGIC = [
    (b'\xff\xd8\xff', 'jpg'),
    (b'\x89PNG\r\n\x1a\n', 'png'),
    (b'GIF87a', 'gif'),
    (b'GIF89a', 'gif'),
    (b'BM', 'bmp'),
]


def detect_ext(data):
    for magic, ext in EXT_BY_MAGIC:
        if data.startswith(magic):
            return ext
    if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return 'webp'
    return None


def optimize_bytes(data, ext):
    """Resize (longest side <= MAX_DIM) and recompress, keeping the format.

    Returns optimized bytes, or the original bytes if PIL can't handle it.
    """
    if ext not in ('jpg', 'png', 'webp'):
        return data
    try:
        img = Image.open(io.BytesIO(data))
        img = ImageOps.exif_transpose(img)  # honour camera orientation
    except Exception:
        return data
    if max(img.size) > MAX_DIM:
        img.thumbnail((MAX_DIM, MAX_DIM), Image.LANCZOS)
    out = io.BytesIO()
    if ext == 'png':
        has_alpha = img.mode in ('RGBA', 'LA', 'P') and (
            img.mode != 'P' or 'transparency' in img.info
        )
        if has_alpha:
            img.convert('RGBA').save(out, format='PNG', optimize=True)
        else:
            img.convert('RGB').save(out, format='JPEG', quality=JPEG_QUALITY, optimize=True, progressive=True)
            return out.getvalue(), 'jpg'  # png without alpha -> jpg
    elif ext == 'webp':
        img.save(out, format='WEBP', quality=JPEG_QUALITY, method=6)
    else:
        img.convert('RGB').save(out, format='JPEG', quality=JPEG_QUALITY, optimize=True, progressive=True)
    result = out.getvalue()
    # Never let optimization make a file bigger.
    return (result if len(result) < len(data) else data)


def normalize_optimize(data, ext):
    """optimize_bytes may switch png->jpg; always return (bytes, ext)."""
    r = optimize_bytes(data, ext)
    if isinstance(r, tuple):
        return r
    return r, ext


def existing_file(base):
    for ext in ('jpg', 'png', 'webp', 'gif', 'bmp'):
        fn = f'{base}.{ext}'
        if os.path.exists(os.path.join(IMG_DIR, fn)):
            return fn
    return None


def fetch_drive(file_id):
    """Download a Drive file, following the large-file confirmation page."""
    url = f'https://drive.google.com/uc?export=download&id={file_id}'
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor())
    opener.addheaders = [('User-Agent', 'Mozilla/5.0')]
    for _ in range(2):
        with opener.open(url, timeout=45) as resp:
            data = resp.read()
            ctype = resp.headers.get('Content-Type', '')
        if 'text/html' in ctype:
            token = re.search(rb'confirm=([\w-]+)', data)
            uuid = re.search(rb'name="uuid" value="([\w-]+)"', data)
            if token:
                url = (f'https://drive.usercontent.google.com/download?id={file_id}'
                       f'&export=download&confirm={token.group(1).decode()}'
                       + (f'&uuid={uuid.group(1).decode()}' if uuid else ''))
                continue
            raise RuntimeError('Drive returned an HTML page (no confirm token)')
        return data
    raise RuntimeError('Drive confirmation loop exhausted')


def download_one(base, file_id):
    """Return (base, filename, status). status: 'skip' | 'ok' | 'fail:<reason>'."""
    have = existing_file(base)
    if have:
        return base, have, 'skip'
    try:
        data = fetch_drive(file_id)
    except (urllib.error.URLError, RuntimeError, TimeoutError) as e:
        return base, None, f'fail:{type(e).__name__}'
    ext = detect_ext(data)
    if not ext:
        return base, None, 'fail:not-an-image'
    data, ext = normalize_optimize(data, ext)
    fn = f'{base}.{ext}'
    with open(os.path.join(IMG_DIR, fn), 'wb') as f:
        f.write(data)
    return base, fn, 'ok'


def optimize_existing():
    """One-time pass: shrink already-downloaded catalog images in place."""
    before = after = 0
    changed = 0
    files = [f for f in os.listdir(IMG_DIR)
             if re.match(r'^(group-|[0-9]{3}[.\-])', f)
             and f.rsplit('.', 1)[-1].lower() in ('jpg', 'png', 'webp')]
    for i, f in enumerate(files, 1):
        path = os.path.join(IMG_DIR, f)
        ext = f.rsplit('.', 1)[-1].lower()
        with open(path, 'rb') as fh:
            data = fh.read()
        before += len(data)
        new, new_ext = normalize_optimize(data, 'jpg' if ext == 'jpg' else ext)
        if len(new) < len(data):
            os.remove(path)
            base = f.rsplit('.', 1)[0]
            with open(os.path.join(IMG_DIR, f'{base}.{new_ext}'), 'wb') as fh:
                fh.write(new)
            after += len(new)
            changed += 1
        else:
            after += len(data)
        if i % 100 == 0 or i == len(files):
            print(f'  optimized {i}/{len(files)} (shrunk {changed})')
    print(f'Optimize: {round(before/1e6,1)}MB -> {round(after/1e6,1)}MB across {len(files)} files')


def download_images(categories, do_download):
    os.makedirs(IMG_DIR, exist_ok=True)
    # Collect base -> drive id (products + group covers).
    jobs = {}
    prod_by_base = {}
    group_by_base = {}
    for cat in categories:
        for grp in cat['groups'].values():
            if grp['imageId']:
                b = f"group-{cat['slug']}-{grp['slug']}"
                jobs[b] = grp['imageId']
                group_by_base[b] = grp
            for p in grp['products']:
                if p['imageId']:
                    jobs[p['id']] = p['imageId']
                    prod_by_base[p['id']] = p

    results = {}
    if do_download:
        counts = {'ok': 0, 'skip': 0, 'fail': 0}
        with ThreadPoolExecutor(max_workers=12) as ex:
            futs = {ex.submit(download_one, b, fid): b for b, fid in jobs.items()}
            done = 0
            total = len(futs)
            for fut in as_completed(futs):
                base, fn, status = fut.result()
                done += 1
                if fn:
                    results[base] = fn
                key = 'ok' if status == 'ok' else 'skip' if status == 'skip' else 'fail'
                counts[key] += 1
                if status.startswith('fail'):
                    print(f'  [FAIL] {base}: {status}', file=sys.stderr)
                if done % 50 == 0 or done == total:
                    print(f'  ...{done}/{total} (ok={counts["ok"]} skip={counts["skip"]} fail={counts["fail"]})')
        print(f'Images: ok={counts["ok"]} skip={counts["skip"]} fail={counts["fail"]}')
    else:
        # No download: map to whatever already exists on disk.
        for b in jobs:
            fn = existing_file(b)
            if fn:
                results[b] = fn

    # Wire resolved filenames back onto the data as public paths.
    for base, p in prod_by_base.items():
        if base in results:
            p['image'] = f'/product/{results[base]}'
    for base, grp in group_by_base.items():
        if base in results:
            grp['image'] = f'/product/{results[base]}'


# --------------------------------------------------------------------------- #
# Emit TypeScript
# --------------------------------------------------------------------------- #
def s(v):
    return json.dumps(v, ensure_ascii=False)


def arr(vs):
    return '[]' if not vs else '[' + ', '.join(s(x) for x in vs) + ']'


def emit_ts(categories):
    L = []
    w = L.append
    w("// AUTO-GENERATED by `scripts/build_products.py` from")
    w("// `public/260717 _ Danh sách sản phẩm An Thái.xlsx`.")
    w("// Nguồn: 2 sheet `VN AT - XP - MAT` + `VN XCBB`. Cấu trúc 3 cấp:")
    w("// Bộ phận (category) → Nhóm mẹ (group) → Sản phẩm con (product).")
    w("// Đừng sửa tay — chạy lại script khi dữ liệu Excel thay đổi.")
    w("")
    w("export interface Product {")
    for f in ('id', 'code', 'name', 'fullName', 'brand', 'unit', 'vehicles'):
        w(f"  {f}: string")
    w("  specs: string[]")
    w("  intro: string[]")
    w("  categorySlug: string")
    w("  groupSlug: string")
    w("  /** Public path under /public, or '' when no image was provided. */")
    w("  image: string")
    w("}")
    w("")
    w("export interface ProductGroup {")
    w("  slug: string")
    w("  name: string")
    w("  categorySlug: string")
    w("  image: string")
    w("  products: Product[]")
    w("}")
    w("")
    w("export interface ProductCategory {")
    w("  slug: string")
    w("  name: string")
    w("  groups: ProductGroup[]")
    w("}")
    w("")
    w("export const productCategories: ProductCategory[] = [")
    for cat in categories:
        w("  {")
        w(f"    slug: {s(cat['slug'])},")
        w(f"    name: {s(cat['name'])},")
        w("    groups: [")
        for grp in cat['groups'].values():
            w("      {")
            w(f"        slug: {s(grp['slug'])},")
            w(f"        name: {s(grp['name'])},")
            w(f"        categorySlug: {s(cat['slug'])},")
            w(f"        image: {s(grp['image'])},")
            w("        products: [")
            for p in grp['products']:
                w("          {")
                w(f"            id: {s(p['id'])},")
                w(f"            code: {s(p['code'])},")
                w(f"            name: {s(p['name'])},")
                w(f"            fullName: {s(p['fullName'])},")
                w(f"            brand: {s(p['brand'])},")
                w(f"            unit: {s(p['unit'])},")
                w(f"            vehicles: {s(p['vehicles'])},")
                w(f"            specs: {arr(p['specs'])},")
                w(f"            intro: {arr(p['intro'])},")
                w(f"            categorySlug: {s(p['categorySlug'])},")
                w(f"            groupSlug: {s(p['groupSlug'])},")
                w(f"            image: {s(p['image'])},")
                w("          },")
            w("        ],")
            w("      },")
        w("    ],")
        w("  },")
    w("]")
    w("")
    # Brand list in display order.
    brands = []
    for cat in categories:
        for grp in cat['groups'].values():
            for p in grp['products']:
                if p['brand'] not in brands:
                    brands.append(p['brand'])
    brands = [b for b in BRAND_ORDER if b in brands] + [b for b in brands if b not in BRAND_ORDER]
    w("// Danh sách thương hiệu (theo thứ tự hiển thị) — feed cho brand filter.")
    w(f"export const productBrands: string[] = {arr(brands)}")
    w("")
    w("// Phẳng hóa toàn bộ sản phẩm — dùng cho search & tra cứu nhanh.")
    w("export const allProducts: Product[] = productCategories.flatMap((category) =>")
    w("  category.groups.flatMap((group) => group.products),")
    w(")")
    w("")
    w("// Tra cứu O(1) theo id sản phẩm.")
    w("export const productsById: Record<string, Product> = Object.fromEntries(")
    w("  allProducts.map((product) => [product.id, product]),")
    w(")")
    w("")
    with open(TS_OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L) + '\n')
    total = sum(len(g['products']) for c in categories for g in c['groups'].values())
    with_img = sum(1 for c in categories for g in c['groups'].values() for p in g['products'] if p['image'])
    print(f'Wrote {TS_OUT}: {total} products ({with_img} with image), '
          f'{sum(len(c["groups"]) for c in categories)} groups.')


def main():
    if '--optimize-existing' in sys.argv:
        optimize_existing()
        if '--no-download' in sys.argv and len(sys.argv) <= 3:
            # optimize-only run still needs the TS to pick up any png->jpg renames
            pass
    do_download = '--no-download' not in sys.argv
    categories = extract()
    download_images(categories, do_download)
    emit_ts(categories)


if __name__ == '__main__':
    main()
